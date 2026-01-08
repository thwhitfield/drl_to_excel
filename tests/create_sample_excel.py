"""
Create sample Excel decision tables for testing.

This script generates Excel files in the standard Drools decision table format.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def create_fraud_rules_excel(output_path: str | Path):
    """
    Create a sample fraud detection decision table.

    Format:
        Row 1: RuleSet | com.example.fraud | Import | com.example.model.Transaction
        Row 2: RuleTable FraudRules
        Row 3: CONDITION | CONDITION | CONDITION | ACTION
        Row 4: $tx : Transaction | $tx : Transaction | $tx : Transaction | $result : Result
        Row 5: score > $1 | amount >= $1, amount < $2 | category == $1 | decision = "$1"
        Row 6: Score Threshold | Amount Range | Category | Decision
        Row 7+: Data rows
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "FraudRules"

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    condition_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    action_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    # Row 1: RuleSet and Import
    ws["A1"] = "RuleSet"
    ws["B1"] = "com.example.fraud"
    ws["C1"] = "Import"
    ws["D1"] = "com.example.model.Transaction"
    ws["E1"] = "com.example.model.Result"

    # Row 2: RuleTable
    ws["A2"] = "RuleTable FraudRules"

    # Row 3: CONDITION / ACTION markers
    ws["A3"] = "CONDITION"
    ws["B3"] = "CONDITION"
    ws["C3"] = "CONDITION"
    ws["D3"] = "ACTION"

    # Row 4: Fact patterns
    ws["A4"] = "$tx : Transaction"
    ws["B4"] = "$tx : Transaction"
    ws["C4"] = "$tx : Transaction"
    ws["D4"] = "$result : Result"

    # Row 5: Templates
    ws["A5"] = "score > $1"
    ws["B5"] = "amount >= $1, amount < $2"
    ws["C5"] = 'category == $1'
    ws["D5"] = 'decision = "$1"'

    # Row 6: Column labels
    ws["A6"] = "Score Threshold"
    ws["B6"] = "Amount Range"
    ws["C6"] = "Category"
    ws["D6"] = "Decision"

    # Apply header styling
    for row in range(1, 7):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            if row >= 3:
                if col <= 3:
                    cell.fill = condition_fill
                else:
                    cell.fill = action_fill

    # Data rows
    data = [
        # High score, high amount → DECLINE
        (0.8, "1000, 10000", "ELECTRONICS", "DECLINE"),
        # High score, any amount → REVIEW
        (0.7, "", "", "REVIEW"),
        # Medium score, high amount electronics → REVIEW
        (0.5, "500, 5000", "ELECTRONICS", "REVIEW"),
        # Low score, small amount → APPROVE
        ("", "0, 100", "", "APPROVE"),
        # Category-specific rule
        ("", "", "GAMBLING", "DECLINE"),
    ]

    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value if value else None)

    # Adjust column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12

    # Save
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Created: {output_path}")


def create_simple_score_rules_excel(output_path: str | Path):
    """
    Create a simpler decision table with just score thresholds.

    This is useful for testing basic parsing without range conditions.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ScoreRules"

    # Row 1: RuleSet
    ws["A1"] = "RuleSet"
    ws["B1"] = "com.example.scoring"

    # Row 2: RuleTable
    ws["A2"] = "RuleTable ScoreRules"

    # Row 3: CONDITION / ACTION
    ws["A3"] = "CONDITION"
    ws["B3"] = "ACTION"

    # Row 4: Fact patterns
    ws["A4"] = "$tx : Transaction"
    ws["B4"] = "$result : Result"

    # Row 5: Templates
    ws["A5"] = "score >= $1"
    ws["B5"] = 'decision = "$1"'

    # Row 6: Labels
    ws["A6"] = "Min Score"
    ws["B6"] = "Decision"

    # Data rows
    data = [
        (0.9, "DECLINE"),
        (0.7, "REVIEW"),
        (0.5, "REVIEW"),
        (0.0, "APPROVE"),
    ]

    for row_idx, row_data in enumerate(data, start=7):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    # Create sample files in tests/fixtures directory
    fixtures_dir = Path(__file__).parent / "fixtures"
    fixtures_dir.mkdir(exist_ok=True)

    create_fraud_rules_excel(fixtures_dir / "fraud_rules.xlsx")
    create_simple_score_rules_excel(fixtures_dir / "score_rules.xlsx")

    print("\nSample Excel files created successfully!")
