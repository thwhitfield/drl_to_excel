"""
Excel Decision Table Writer.

Writes IR (RuleSet) to Excel files in the standard Drools decision table format.

Standard Drools Excel Format:
    Row 1: RuleSet, package name | Import, class1, class2...
    Row 2: RuleTable <TableName>
    Row 3: CONDITION | CONDITION | ... | ACTION | ACTION
    Row 4: $tx : Transaction | $tx : Transaction | ... | $result : Result
    Row 5: score > $1 | amount >= $1, amount < $2 | ... | decision = "$1"
    Row 6: Score Threshold | Amount Range | ... | Decision
    Row 7+: Data rows
"""

from pathlib import Path
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from drl_to_excel.ir import (
    RuleSet, Rule, FactPattern, Action, Condition,
    SimpleCondition, RangeCondition, BucketCondition,
    Operator, ActionType,
)


@dataclass
class ColumnSpec:
    """Specification for a decision table column."""
    column_type: str  # "CONDITION" or "ACTION"
    fact_pattern: str  # e.g., "$tx : Transaction"
    template: str  # e.g., "score > $1"
    label: str  # Human-readable column name
    binding: str | None = None
    fact_type: str | None = None
    field: str | None = None


@dataclass
class ExcelWriterConfig:
    """Configuration for Excel output formatting."""
    header_font: Font = field(default_factory=lambda: Font(bold=True))
    condition_fill: PatternFill = field(
        default_factory=lambda: PatternFill(
            start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"
        )
    )
    action_fill: PatternFill = field(
        default_factory=lambda: PatternFill(
            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
        )
    )
    header_fill: PatternFill = field(
        default_factory=lambda: PatternFill(
            start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
        )
    )
    column_width: int = 18


class ExcelWriter:
    """Writer for Drools decision tables in Excel format."""

    def __init__(self, config: ExcelWriterConfig | None = None):
        self.config = config or ExcelWriterConfig()
        self.columns: list[ColumnSpec] = []

    def write(self, ruleset: RuleSet, file_path: str | Path) -> None:
        """Write a RuleSet to an Excel file."""
        # Analyze the ruleset to determine column structure
        self._analyze_ruleset(ruleset)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = ruleset.rule_table_name or ruleset.name or "Rules"

        # Write header rows
        self._write_headers(ws, ruleset)

        # Write data rows
        self._write_data_rows(ws, ruleset)

        # Apply formatting
        self._apply_formatting(ws)

        # Save
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    def _analyze_ruleset(self, ruleset: RuleSet) -> None:
        """Analyze rules to determine column structure."""
        self.columns = []

        # Collect all unique conditions and actions across rules
        condition_specs: dict[str, ColumnSpec] = {}
        action_specs: dict[str, ColumnSpec] = {}

        for rule in ruleset.rules:
            # Analyze conditions
            for pattern in rule.fact_patterns:
                for condition in pattern.conditions:
                    key = self._get_condition_key(condition, pattern)
                    if key not in condition_specs:
                        condition_specs[key] = self._create_condition_spec(
                            condition, pattern
                        )

            # Analyze actions
            for action in rule.actions:
                key = self._get_action_key(action)
                if key not in action_specs:
                    action_specs[key] = self._create_action_spec(action)

        # Order: conditions first, then actions
        self.columns = list(condition_specs.values()) + list(action_specs.values())

    def _get_condition_key(self, condition: Condition, pattern: FactPattern) -> str:
        """Generate a unique key for a condition column."""
        field = condition.get_field_name()
        if isinstance(condition, RangeCondition):
            return f"{pattern.binding}:{pattern.fact_type}:{field}:range"
        elif isinstance(condition, BucketCondition):
            return f"{pattern.binding}:{pattern.fact_type}:{field}:bucket"
        else:
            return f"{pattern.binding}:{pattern.fact_type}:{field}:{condition.operator.value}"

    def _get_action_key(self, action: Action) -> str:
        """Generate a unique key for an action column."""
        return f"{action.binding}:{action.action_type.value}:{action.target}"

    def _create_condition_spec(
        self,
        condition: Condition,
        pattern: FactPattern
    ) -> ColumnSpec:
        """Create a column specification for a condition."""
        field = condition.get_field_name()
        fact_pattern_str = f"${pattern.binding} : {pattern.fact_type}"

        if isinstance(condition, RangeCondition):
            template = f"{field} >= $1, {field} < $2"
            label = f"{self._humanize(field)} Range"
        elif isinstance(condition, BucketCondition):
            template = f"({field} % {condition.modulo}) in ($1)"
            label = f"{self._humanize(field)} Bucket"
        elif isinstance(condition, SimpleCondition):
            op = condition.operator.value
            template = f"{field} {op} $1"
            label = self._humanize(field)
        else:
            template = f"{field} $1"
            label = self._humanize(field)

        return ColumnSpec(
            column_type="CONDITION",
            fact_pattern=fact_pattern_str,
            template=template,
            label=label,
            binding=pattern.binding,
            fact_type=pattern.fact_type,
            field=field,
        )

    def _create_action_spec(self, action: Action) -> ColumnSpec:
        """Create a column specification for an action."""
        if action.action_type == ActionType.SET_FIELD:
            fact_pattern = f"${action.binding} : Result" if action.binding else ""
            template = f'{action.target} = "$1"'
            label = self._humanize(action.target)
        elif action.action_type == ActionType.INSERT_FACT:
            fact_pattern = ""
            template = f"insert(new {action.target}($1))"
            label = f"Insert {action.target}"
        else:
            fact_pattern = f"${action.binding}" if action.binding else ""
            template = str(action.action_type.value)
            label = action.action_type.value.title()

        return ColumnSpec(
            column_type="ACTION",
            fact_pattern=fact_pattern,
            template=template,
            label=label,
            binding=action.binding,
        )

    def _humanize(self, name: str) -> str:
        """Convert camelCase or snake_case to Title Case."""
        # Handle camelCase
        import re
        name = re.sub(r'([a-z])([A-Z])', r'\1 \2', name)
        # Handle snake_case
        name = name.replace('_', ' ')
        return name.title()

    def _write_headers(self, ws, ruleset: RuleSet) -> None:
        """Write the header rows (1-6)."""
        # Row 1: RuleSet and Imports
        ws["A1"] = "RuleSet"
        ws["B1"] = ruleset.package

        if ruleset.imports:
            ws["C1"] = "Import"
            for idx, imp in enumerate(ruleset.imports):
                ws.cell(row=1, column=4 + idx, value=imp)

        # Row 2: RuleTable
        table_name = ruleset.rule_table_name or ruleset.name or "Rules"
        ws["A2"] = f"RuleTable {table_name}"

        # Rows 3-6: Column definitions
        for col_idx, col_spec in enumerate(self.columns, start=1):
            # Row 3: CONDITION / ACTION
            ws.cell(row=3, column=col_idx, value=col_spec.column_type)

            # Row 4: Fact pattern
            ws.cell(row=4, column=col_idx, value=col_spec.fact_pattern)

            # Row 5: Template
            ws.cell(row=5, column=col_idx, value=col_spec.template)

            # Row 6: Label
            ws.cell(row=6, column=col_idx, value=col_spec.label)

    def _write_data_rows(self, ws, ruleset: RuleSet) -> None:
        """Write data rows for each rule."""
        for row_idx, rule in enumerate(ruleset.rules, start=7):
            self._write_rule_row(ws, row_idx, rule)

    def _write_rule_row(self, ws, row: int, rule: Rule) -> None:
        """Write a single rule as a data row."""
        for col_idx, col_spec in enumerate(self.columns, start=1):
            value = self._get_cell_value(rule, col_spec)
            if value is not None:
                ws.cell(row=row, column=col_idx, value=value)

    def _get_cell_value(self, rule: Rule, col_spec: ColumnSpec) -> str | None:
        """Get the cell value for a rule/column combination."""
        if col_spec.column_type == "CONDITION":
            return self._get_condition_value(rule, col_spec)
        else:
            return self._get_action_value(rule, col_spec)

    def _get_condition_value(self, rule: Rule, col_spec: ColumnSpec) -> str | None:
        """Extract condition value for a column."""
        for pattern in rule.fact_patterns:
            if pattern.binding != col_spec.binding:
                continue
            if pattern.fact_type != col_spec.fact_type:
                continue

            for condition in pattern.conditions:
                if condition.get_field_name() != col_spec.field:
                    continue

                if isinstance(condition, RangeCondition):
                    parts = []
                    if condition.min_value is not None:
                        parts.append(str(condition.min_value))
                    if condition.max_value is not None:
                        parts.append(str(condition.max_value))
                    return ", ".join(parts) if parts else None

                elif isinstance(condition, BucketCondition):
                    return ", ".join(str(v) for v in condition.bucket_values)

                elif isinstance(condition, SimpleCondition):
                    return self._format_value(condition.value)

        return None

    def _get_action_value(self, rule: Rule, col_spec: ColumnSpec) -> str | None:
        """Extract action value for a column."""
        for action in rule.actions:
            if action.binding != col_spec.binding:
                continue
            if action.target != col_spec.label.lower().replace(" ", ""):
                # Try matching on target field
                if action.target.lower() != col_spec.label.lower().replace(" ", ""):
                    # Check if it's a field match
                    if col_spec.field and action.target.lower() == col_spec.field.lower():
                        pass
                    elif action.target.lower() == self._humanize(col_spec.label).lower().replace(" ", ""):
                        pass
                    else:
                        continue

            return self._format_value(action.value)

        # Fallback: match by action type and target
        for action in rule.actions:
            if action.action_type == ActionType.SET_FIELD:
                target_label = self._humanize(action.target)
                if target_label == col_spec.label:
                    return self._format_value(action.value)

        return None

    def _format_value(self, value) -> str:
        """Format a value for Excel cell."""
        if value is None:
            return ""
        if isinstance(value, bool):
            return str(value).lower()
        if isinstance(value, (int, float)):
            return value  # Keep as number for Excel
        return str(value)

    def _apply_formatting(self, ws) -> None:
        """Apply styling to the worksheet."""
        # Apply header formatting to rows 1-6
        for row in range(1, 7):
            for col in range(1, len(self.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.config.header_font

                if row >= 3:
                    col_spec = self.columns[col - 1] if col <= len(self.columns) else None
                    if col_spec:
                        if col_spec.column_type == "CONDITION":
                            cell.fill = self.config.condition_fill
                        else:
                            cell.fill = self.config.action_fill

        # Row 1-2 get header fill
        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).fill = self.config.header_fill
            ws.cell(row=2, column=col).fill = self.config.header_fill

        # Adjust column widths
        for col in range(1, len(self.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = self.config.column_width


def write_excel(ruleset: RuleSet, file_path: str | Path) -> None:
    """Convenience function to write a RuleSet to Excel."""
    writer = ExcelWriter()
    writer.write(ruleset, file_path)
