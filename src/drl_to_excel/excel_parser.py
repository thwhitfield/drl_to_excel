"""
Excel Decision Table Parser.

Parses standard Drools decision table format from Excel files into IR.

Standard Drools Excel Format:
    Row 1: RuleSet, package name | Import, class1, class2...
    Row 2: RuleTable <TableName>
    Row 3: CONDITION | CONDITION | ... | ACTION | ACTION
    Row 4: $tx : Transaction | $tx : Transaction | ... | $result : Result
    Row 5: score $1 | amount >= $1, amount < $2 | ... | decision = "$1"
    Row 6: Score Threshold | Amount Range | ... | Decision
    Row 7+: Data rows (0.8 | 100, 500 | ... | DECLINE)
"""

import re
from pathlib import Path
from dataclasses import dataclass
from typing import Any
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from drl_to_excel.ir import (
    Operator, ActionType,
    SimpleCondition, RangeCondition, Condition,
    FactPattern, Action, Rule, RuleSet,
)


@dataclass
class ColumnDefinition:
    """Parsed column definition from header rows."""
    col_index: int
    column_type: str  # "CONDITION" or "ACTION"
    fact_pattern: str  # e.g., "$tx : Transaction"
    template: str  # e.g., "score > $1" or "amount >= $1, amount < $2"
    label: str  # Human-readable column name

    # Parsed from fact_pattern
    binding: str | None = None
    fact_type: str | None = None


@dataclass
class ParsedTemplate:
    """Parsed template with placeholders identified."""
    field: str
    operator: str
    placeholder_count: int
    is_range: bool = False
    raw_template: str = ""


class ExcelParserError(Exception):
    """Raised when Excel parsing fails."""
    pass


class ExcelParser:
    """Parser for Drools decision tables in Excel format."""

    # Standard Drools decision table keywords
    KEYWORDS = {
        "RuleSet": "RULESET",
        "Import": "IMPORT",
        "RuleTable": "RULETABLE",
        "CONDITION": "CONDITION",
        "ACTION": "ACTION",
    }

    # Operator mapping for template parsing
    OPERATOR_MAP = {
        "==": Operator.EQ,
        "!=": Operator.NE,
        ">": Operator.GT,
        ">=": Operator.GE,
        "<": Operator.LT,
        "<=": Operator.LE,
        "in": Operator.IN,
        "not in": Operator.NOT_IN,
        "matches": Operator.MATCHES,
        "contains": Operator.CONTAINS,
    }

    def __init__(self):
        self.columns: list[ColumnDefinition] = []
        self.ruleset_name: str = ""
        self.package: str = ""
        self.imports: list[str] = []
        self.rule_table_name: str = ""

    def parse_file(self, file_path: str | Path) -> RuleSet:
        """Parse an Excel decision table file into IR."""
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        if sheet is None:
            raise ExcelParserError("No active worksheet found")
        return self.parse_sheet(sheet)

    def parse_sheet(self, sheet: Worksheet) -> RuleSet:
        """Parse a worksheet into a RuleSet."""
        self._reset()

        # Find the decision table start
        start_row = self._find_table_start(sheet)

        # Parse header rows
        self._parse_headers(sheet, start_row)

        # Parse data rows into rules
        # Header rows: 1=RuleSet, 2=RuleTable, 3=CONDITION/ACTION, 4=Pattern, 5=Template, 6=Label
        # Data starts at row 7 (start_row + 6)
        rules = self._parse_data_rows(sheet, start_row + 6)

        return RuleSet(
            name=self.ruleset_name or self.rule_table_name,
            package=self.package,
            imports=self.imports,
            rules=rules,
            rule_table_name=self.rule_table_name,
        )

    def _reset(self):
        """Reset parser state."""
        self.columns = []
        self.ruleset_name = ""
        self.package = ""
        self.imports = []
        self.rule_table_name = ""

    def _find_table_start(self, sheet: Worksheet) -> int:
        """Find the row where the decision table starts (RuleSet row)."""
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
            for cell in row:
                if cell and str(cell).strip().upper() == "RULESET":
                    return row_idx
        raise ExcelParserError("Could not find RuleSet header in first 20 rows")

    def _parse_headers(self, sheet: Worksheet, start_row: int):
        """Parse the 5 header rows of the decision table."""
        # Row 1: RuleSet and Import
        self._parse_ruleset_row(sheet, start_row)

        # Row 2: RuleTable
        self._parse_ruletable_row(sheet, start_row + 1)

        # Rows 3-5: Column definitions
        self._parse_column_definitions(sheet, start_row + 2)

    def _parse_ruleset_row(self, sheet: Worksheet, row: int):
        """Parse RuleSet, package, and imports from row 1."""
        row_values = list(sheet.iter_rows(min_row=row, max_row=row, values_only=True))[0]

        for idx, cell in enumerate(row_values):
            if cell is None:
                continue
            cell_str = str(cell).strip()

            if cell_str.upper() == "RULESET":
                # Next cell should be the package/name
                if idx + 1 < len(row_values) and row_values[idx + 1]:
                    self.package = str(row_values[idx + 1]).strip()
                    self.ruleset_name = self.package.split(".")[-1]

            elif cell_str.upper() == "IMPORT":
                # Collect all following non-empty cells as imports
                for import_cell in row_values[idx + 1:]:
                    if import_cell:
                        self.imports.append(str(import_cell).strip())
                break

    def _parse_ruletable_row(self, sheet: Worksheet, row: int):
        """Parse RuleTable name from row 2."""
        row_values = list(sheet.iter_rows(min_row=row, max_row=row, values_only=True))[0]

        for cell in row_values:
            if cell is None:
                continue
            cell_str = str(cell).strip()

            if cell_str.upper().startswith("RULETABLE"):
                # Extract table name after "RuleTable"
                match = re.match(r"RULETABLE\s+(.+)", cell_str, re.IGNORECASE)
                if match:
                    self.rule_table_name = match.group(1).strip()
                break

    def _parse_column_definitions(self, sheet: Worksheet, start_row: int):
        """Parse column definitions from rows 3-5 (type, pattern, template, label)."""
        # Get the three rows
        rows = list(sheet.iter_rows(
            min_row=start_row,
            max_row=start_row + 3,
            values_only=True
        ))

        if len(rows) < 4:
            raise ExcelParserError("Not enough header rows for column definitions")

        type_row = rows[0]       # CONDITION / ACTION
        pattern_row = rows[1]    # $tx : Transaction
        template_row = rows[2]   # score > $1
        label_row = rows[3]      # Score Threshold

        # Find all columns with definitions
        for col_idx, col_type in enumerate(type_row):
            if col_type is None:
                continue

            col_type_str = str(col_type).strip().upper()
            if col_type_str not in ("CONDITION", "ACTION"):
                continue

            # Get the other row values for this column
            pattern = str(pattern_row[col_idx]).strip() if col_idx < len(pattern_row) and pattern_row[col_idx] else ""
            template = str(template_row[col_idx]).strip() if col_idx < len(template_row) and template_row[col_idx] else ""
            label = str(label_row[col_idx]).strip() if col_idx < len(label_row) and label_row[col_idx] else f"Column{col_idx}"

            col_def = ColumnDefinition(
                col_index=col_idx,
                column_type=col_type_str,
                fact_pattern=pattern,
                template=template,
                label=label,
            )

            # Parse binding and fact type from pattern
            self._parse_fact_pattern(col_def)

            self.columns.append(col_def)

    def _parse_fact_pattern(self, col_def: ColumnDefinition):
        """Extract binding and fact type from pattern like '$tx : Transaction'."""
        pattern = col_def.fact_pattern
        match = re.match(r"\$(\w+)\s*:\s*(\w+)", pattern)
        if match:
            col_def.binding = match.group(1)
            col_def.fact_type = match.group(2)

    def _parse_data_rows(self, sheet: Worksheet, start_row: int) -> list[Rule]:
        """Parse data rows into Rule objects."""
        rules = []
        rule_number = 1

        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            # Check if row has any data
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            rule = self._parse_data_row(row, rule_number)
            if rule:
                rules.append(rule)
                rule_number += 1

        return rules

    def _parse_data_row(self, row: tuple, rule_number: int) -> Rule | None:
        """Parse a single data row into a Rule."""
        # Group conditions by fact type
        fact_patterns: dict[str, FactPattern] = {}
        actions: list[Action] = []

        for col_def in self.columns:
            if col_def.col_index >= len(row):
                continue

            cell_value = row[col_def.col_index]
            if cell_value is None or str(cell_value).strip() == "":
                continue

            cell_str = str(cell_value).strip()

            if col_def.column_type == "CONDITION":
                condition = self._parse_condition(col_def, cell_str)
                if condition:
                    # Get or create fact pattern
                    key = f"{col_def.binding}:{col_def.fact_type}"
                    if key not in fact_patterns:
                        fact_patterns[key] = FactPattern(
                            fact_type=col_def.fact_type or "Object",
                            binding=col_def.binding,
                            conditions=[],
                        )
                    fact_patterns[key].conditions.append(condition)

            elif col_def.column_type == "ACTION":
                action = self._parse_action(col_def, cell_str)
                if action:
                    actions.append(action)

        # Only create rule if we have conditions or actions
        if not fact_patterns and not actions:
            return None

        return Rule(
            name=f"{self.rule_table_name}_{rule_number}",
            fact_patterns=list(fact_patterns.values()),
            actions=actions,
        )

    def _parse_condition(self, col_def: ColumnDefinition, value: str) -> Condition | None:
        """Parse a condition from template and cell value."""
        template = col_def.template

        # Check for range conditions (multiple comparisons)
        if "," in template and "$1" in template and "$2" in template:
            return self._parse_range_condition(template, value)

        # Parse simple condition
        return self._parse_simple_condition(template, value)

    def _parse_simple_condition(self, template: str, value: str) -> SimpleCondition | None:
        """Parse a simple condition like 'score > $1' with value '0.8'."""
        # Pattern: field operator $n
        # Also handle: field $1 (implicit ==)
        pattern = r"(\w+)\s*(==|!=|>=|<=|>|<|in|not in|matches|contains)?\s*\$\d+"

        match = re.match(pattern, template.strip())
        if not match:
            # Try without operator (implicit equality)
            match = re.match(r"(\w+)\s*\$\d+", template.strip())
            if match:
                field = match.group(1)
                operator = Operator.EQ
            else:
                return None
        else:
            field = match.group(1)
            op_str = match.group(2) if match.group(2) else "=="
            operator = self.OPERATOR_MAP.get(op_str, Operator.EQ)

        # Convert value to appropriate type
        typed_value = self._convert_value(value)

        return SimpleCondition(field=field, operator=operator, value=typed_value)

    def _parse_range_condition(self, template: str, value: str) -> RangeCondition | None:
        """Parse a range condition like 'amount >= $1, amount < $2' with value '100, 500'."""
        # Split template and value by comma
        template_parts = [p.strip() for p in template.split(",")]
        value_parts = [p.strip() for p in value.split(",")]

        if len(template_parts) != 2 or len(value_parts) != 2:
            return None

        # Parse field name from first part
        field_match = re.match(r"(\w+)", template_parts[0])
        if not field_match:
            return None
        field = field_match.group(1)

        # Determine operators and values
        min_value = None
        max_value = None
        min_inclusive = True
        max_inclusive = False

        for tpl, val in zip(template_parts, value_parts):
            if ">=" in tpl:
                min_value = self._convert_value(val)
                min_inclusive = True
            elif ">" in tpl:
                min_value = self._convert_value(val)
                min_inclusive = False
            elif "<=" in tpl:
                max_value = self._convert_value(val)
                max_inclusive = True
            elif "<" in tpl:
                max_value = self._convert_value(val)
                max_inclusive = False

        return RangeCondition(
            field=field,
            min_value=min_value,
            max_value=max_value,
            min_inclusive=min_inclusive,
            max_inclusive=max_inclusive,
        )

    def _parse_action(self, col_def: ColumnDefinition, value: str) -> Action | None:
        """Parse an action from template and cell value."""
        template = col_def.template

        # Pattern: field = "$1" or field = $1
        match = re.match(r"(\w+)\s*=\s*\"?\$\d+\"?", template.strip())
        if match:
            field = match.group(1)
            return Action(
                action_type=ActionType.SET_FIELD,
                target=field,
                value=value,
                binding=col_def.binding,
            )

        # Pattern: insert(new FactType($1))
        match = re.match(r"insert\s*\(\s*new\s+(\w+)\s*\(\s*\$\d+\s*\)\s*\)", template.strip())
        if match:
            fact_type = match.group(1)
            return Action(
                action_type=ActionType.INSERT_FACT,
                target=fact_type,
                value=value,
            )

        # Fallback: treat as custom action
        return Action(
            action_type=ActionType.CUSTOM,
            target=col_def.label,
            value=value,
        )

    def _convert_value(self, value: Any) -> Any:
        """Convert value to appropriate Python type."""
        # If already a number, return as-is
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return value

        # If not a string, return as-is
        if not isinstance(value, str):
            return value

        value = value.strip()

        # Boolean
        if value.lower() == "true":
            return True
        if value.lower() == "false":
            return False

        # Numeric
        try:
            if "." in value:
                return float(value)
            return int(value)
        except ValueError:
            pass

        # String (strip quotes if present)
        if (value.startswith('"') and value.endswith('"')) or \
           (value.startswith("'") and value.endswith("'")):
            return value[1:-1]

        return value


def parse_excel(file_path: str | Path) -> RuleSet:
    """Convenience function to parse an Excel decision table."""
    parser = ExcelParser()
    return parser.parse_file(file_path)
